from __future__ import annotations

import calendar
import math
import tempfile
from urllib.error import HTTPError, URLError
import urllib.request
from pathlib import Path

import openpyxl


SOURCE_URL = "https://www.finra.org/sites/default/files/2021-03/margin-statistics.xlsx"
OUTPUT_FILE = Path("data/finra-margin-debt-yoy.csv")


def month_end(year_month: str) -> str:
    year, month = [int(part) for part in year_month.split("-")]
    day = calendar.monthrange(year, month)[1]
    return f"{year:04d}-{month:02d}-{day:02d}"


def download_xlsx() -> Path:
    last_error: Exception | None = None
    payload: bytes | None = None

    for attempt in range(1):
        try:
            with urllib.request.urlopen(SOURCE_URL, timeout=20) as response:
                payload = response.read()
            break
        except (HTTPError, URLError, TimeoutError) as error:
            last_error = error

    if payload is None:
        raise RuntimeError(f"FINRA download failed: {last_error}")

    handle = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    handle.write(payload)
    handle.close()
    return Path(handle.name)


def load_debit_balances(xlsx_path: Path) -> list[tuple[str, float]]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]

    if headers[:2] != [
        "Year-Month",
        "Debit Balances in Customers' Securities Margin Accounts",
    ]:
        raise ValueError(f"Unexpected FINRA headers: {headers}")

    rows: list[tuple[str, float]] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        year_month, debit_balance = row[0], row[1]

        if not year_month or debit_balance is None:
            continue

        value = float(str(debit_balance).replace(",", ""))

        if not math.isfinite(value):
            continue

        rows.append((str(year_month), value))

    return sorted(rows, key=lambda item: item[0])


def calculate_yoy(rows: list[tuple[str, float]]) -> list[tuple[str, float]]:
    balances = dict(rows)
    output: list[tuple[str, float]] = []

    for year_month, value in rows:
        year, month = [int(part) for part in year_month.split("-")]
        prior_key = f"{year - 1:04d}-{month:02d}"
        prior_value = balances.get(prior_key)

        if prior_value and prior_value > 0:
            output.append((month_end(year_month), (value / prior_value - 1) * 100))

    return output


def validate(rows: list[tuple[str, float]]) -> None:
    dates = [date for date, _ in rows]

    if not rows:
        raise ValueError("FINRA Margin Debt YoY has no valid observations.")

    if dates != sorted(dates):
        raise ValueError("FINRA Margin Debt YoY dates are not sorted.")

    if len(dates) != len(set(dates)):
        raise ValueError("FINRA Margin Debt YoY contains duplicate dates.")

    if any(not math.isfinite(value) for _, value in rows):
        raise ValueError("FINRA Margin Debt YoY contains invalid values.")


def atomic_write(rows: list[tuple[str, float]]) -> None:
    temp_file = OUTPUT_FILE.with_suffix(".csv.tmp")
    temp_file.write_text(
        "date,value\n" + "\n".join(f"{date},{value:.2f}" for date, value in rows) + "\n",
        encoding="utf-8",
    )
    validate([
        (line.split(",")[0], float(line.split(",")[1]))
        for line in temp_file.read_text(encoding="utf-8").strip().splitlines()[1:]
    ])
    temp_file.replace(OUTPUT_FILE)


def main() -> None:
    xlsx_path = download_xlsx()
    raw_rows = load_debit_balances(xlsx_path)
    yoy_rows = calculate_yoy(raw_rows)
    validate(yoy_rows)
    atomic_write(yoy_rows)

    print("FINRA Margin Debt YoY validation")
    print(f"Source earliest month: {raw_rows[0][0]}")
    print(f"Source latest month: {raw_rows[-1][0]}")
    print(f"YoY earliest date: {yoy_rows[0][0]}")
    print(f"YoY latest date: {yoy_rows[-1][0]}")
    print(f"Valid observations: {len(yoy_rows)}")
    print(f"Duplicate dates: {len(yoy_rows) - len(set(date for date, _ in yoy_rows))}")


if __name__ == "__main__":
    main()
