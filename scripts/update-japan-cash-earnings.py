#!/usr/bin/env python3
"""Update nationwide monthly cash earnings YoY from the official MHLW release table."""

from __future__ import annotations

import csv
import re
import ssl
import tempfile
from datetime import date
from pathlib import Path
from urllib.parse import urljoin
from urllib.request import Request, urlopen

import certifi
from openpyxl import load_workbook


INDEX_URL = "https://www.mhlw.go.jp/toukei/list/30-1a.html?os=jva"
OUTPUT_FILE = Path("data/japan-cash-earnings-yoy.csv")
USER_AGENT = "market-indicators-dashboard/1.0"
TIMEOUT_SECONDS = 30


def download(url: str) -> bytes:
    request = Request(url, headers={"User-Agent": USER_AGENT})
    ssl_context = ssl.create_default_context(cafile=certifi.where())
    with urlopen(request, timeout=TIMEOUT_SECONDS, context=ssl_context) as response:
        if response.status != 200:
            raise RuntimeError(f"MHLW download failed: HTTP {response.status} for {url}")
        return response.read()


def decode_html(payload: bytes) -> str:
    for encoding in ("shift_jis", "utf-8"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise RuntimeError("Could not decode the MHLW release index.")


def latest_preliminary_page(index_html: str) -> str:
    matches = re.findall(r'href="([^"]+/r\d{2}/\d{4}p/\d{4}p\.html)"', index_html)
    if not matches:
        raise RuntimeError("Could not find a preliminary monthly cash-earnings release link.")
    return sorted(set(matches))[-1]


def release_date_from_url(page_url: str) -> str:
    match = re.search(r"/r(\d{2})/(\d{2})(\d{2})p/", page_url)
    if not match:
        raise RuntimeError(f"Could not determine release month from {page_url}")
    reiwa_year, year_suffix, month = map(int, match.groups())
    year = 2018 + reiwa_year
    if year % 100 != year_suffix:
        raise RuntimeError(f"Unexpected MHLW release year in {page_url}")
    return date(year, month, 1).isoformat()


def first_table_url(page_url: str, page_html: str) -> str:
    match = re.search(r'href="([^"]*xls/\d{4}c01p\.xlsx)"', page_html)
    if not match:
        raise RuntimeError("Could not find the official monthly cash-earnings workbook.")
    return urljoin(page_url, match.group(1))


def read_total_cash_earnings_yoy(workbook_bytes: bytes) -> float:
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as temporary_file:
        temporary_file.write(workbook_bytes)
        temporary_file.flush()
        workbook = load_workbook(temporary_file.name, data_only=True, read_only=True)
        worksheet = workbook.active
        for row in worksheet.iter_rows(values_only=True):
            if not row or not isinstance(row[0], str):
                continue
            if row[0].replace(" ", "").replace("　", "") != "調査産業計":
                continue
            value = row[3] if len(row) > 3 else None
            if isinstance(value, (int, float)):
                return float(value)
    raise RuntimeError("Could not find the total-industry cash earnings year-over-year observation.")


def read_existing() -> dict[str, float]:
    if not OUTPUT_FILE.exists():
        return {}
    with OUTPUT_FILE.open(newline="", encoding="utf-8") as handle:
        return {
            row["date"]: float(row["value"])
            for row in csv.DictReader(handle)
            if row.get("date") and row.get("value")
        }


def write_rows(rows: dict[str, float]) -> None:
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    temporary_file = OUTPUT_FILE.with_suffix(".csv.tmp")
    with temporary_file.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(["date", "value"])
        for observation_date, value in sorted(rows.items()):
            writer.writerow([observation_date, f"{value:.1f}"])
    temporary_file.replace(OUTPUT_FILE)


def main() -> None:
    index_html = decode_html(download(INDEX_URL))
    page_url = urljoin(INDEX_URL, latest_preliminary_page(index_html))
    page_html = decode_html(download(page_url))
    observation_date = release_date_from_url(page_url)
    value = read_total_cash_earnings_yoy(download(first_table_url(page_url, page_html)))
    rows = read_existing()
    rows[observation_date] = value
    write_rows(rows)
    print(f"Japan Cash Earnings YoY: {observation_date} = {value:.1f}%")


if __name__ == "__main__":
    main()
